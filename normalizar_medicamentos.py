"""
Script de Normalización del Catálogo de Medicamentos
=====================================================
Analiza la columna 'agrupador' del archivo medicamentos.xls,
detecta inconsistencias (grupos con mezcla de moléculas o concentraciones)
y genera un archivo Excel reclasificado.

Regla de validez: Un agrupador es válido cuando TODOS sus ítems comparten
la misma molécula (principio activo) Y la misma concentración.
"""

import pandas as pd
import re
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# 1. CARGA DE DATOS
# ─────────────────────────────────────────────────────────────────────────────
RUTA_ENTRADA = r'C:\xampp\htdocs\medicamentos_pendientes\medicamentos.xls'
RUTA_SALIDA  = r'C:\xampp\htdocs\medicamentos_pendientes\medicamentos_normalizado.xlsx'

print("=" * 70)
print("NORMALIZADOR DE CATÁLOGO DE MEDICAMENTOS")
print("=" * 70)
print(f"\nCargando: {RUTA_ENTRADA}")

df = pd.read_excel(RUTA_ENTRADA, engine='xlrd')
df = df.copy()
# Asegurar que todas las columnas de texto sean strings limpias
for col in ['agrupador', 'codigo', 'nombre', 'marca', 'atc', 'cums']:
    df[col] = df[col].astype(str).str.strip()

print(f"Total de registros: {len(df):,}")
print(f"Agrupadores únicos: {df['agrupador'].nunique():,}")


# ─────────────────────────────────────────────────────────────────────────────
# 2. EXTRACCIÓN DE MOLÉCULA Y CONCENTRACIÓN
# ─────────────────────────────────────────────────────────────────────────────

# Palabras clave de formas farmacéuticas que separan el principio activo
# del resto del nombre
FORMAS_FARMACEUTICAS = [
    r'\bTABLETA[S]?\b', r'\bTABLET[S]?\b',
    r'\bTAB\b',
    r'\bCAPSULA[S]?\b', r'\bCAPS\b', r'\bCAP\b',
    r'\bSOLUCION\b', r'\bSOL\b',
    r'\bINYECTABLE\b', r'\bINYECCION\b', r'\bINY\b',
    r'\bJARABE\b', r'\bJBE\b',
    r'\bCREMA\b',
    r'\bGEL\b',
    r'\bUNGUENTO\b', r'\bUNG\b',
    r'\bSUSPENSION\b', r'\bSUSP\b',
    r'\bAMPOLLA\b', r'\bAMP\b',
    r'\bSUPOSITORIO\b', r'\bSUPO\b',
    r'\bOFTALMICO\b', r'\bOFT\b',
    r'\bPOLVO\b', r'\bPOLV\b',
    r'\bGOTAS\b', r'\bGOT\b',
    r'\bCOMPRIMIDO[S]?\b', r'\bCOMP\b',
    r'\bOVULO\b',
    r'\bPARCHE\b',
    r'\bAEROSOL\b',
    r'\bINHALADOR\b', r'\bINH\b',
    r'\bSACHET\b', r'\bSAC\b',
    r'\bELIXIR\b',
    r'\bCOLIRIO\b',
    r'\bEMULSION\b',
    r'\bFRASCO\b', r'\bFCO\b',
    r'\bIMPLANTE\b',
    r'\bSISTEMA\b',
    r'\bNEBULIZADOR\b', r'\bNEB\b',
    r'\bVAGINAL\b',
    r'\bRECTAL\b',
    # Rutas de administración
    r'\bINTRAVENOSO\b', r'\bIV\b',
    r'\bSUBCUTANEO\b', r'\bSC\b',
]

# Patrón principal de concentración – captura la primera dosis relevante
# Orden de especificidad: combinaciones > ratio > porcentaje > simple
PATRON_CONC = re.compile(
    r'(?:'
    # Combinaciones: 250+250+65 MG  ó  3.5+25+5+18 %
    r'(\d+(?:[.,]\d+)?(?:\s*\+\s*\d+(?:[.,]\d+)?){1,6})\s*'
    r'(MG|MCG|G|UI|%|MEQ|MMOL|UM|MICROGRAMO|ML|MG/ML|MCG/ML|UI/ML|MG/G)'
    r'|'
    # Ratio: 150MG/5ML  ó  1MG/ML
    r'(\d+(?:[.,]\d+)?)\s*(MG|MCG|G|UI|%|MEQ|MMOL|UM)\s*/\s*(\d+(?:[.,]\d+)?)\s*(ML|MG|G|L)'
    r'|'
    # Simple: 500 MG  ó  0.025%  ó  3%
    r'(\d+(?:[.,]\d+)?)\s*(MG|MCG|G|UI|%|MEQ|MMOL|UM|MICROGRAMO|ML)'
    r')',
    re.IGNORECASE
)


def normalizar_nombre(nombre: str) -> str:
    """Elimina sufijos como *, **, espacios múltiples y normaliza mayúsculas."""
    n = str(nombre).upper().strip()
    n = re.sub(r'\s*\*+\s*$', '', n)      # quitar * al final
    n = re.sub(r'\s+', ' ', n)             # normalizar espacios
    return n.strip()


def extraer_molecula(nombre_norm: str) -> str:
    """
    Extrae la molécula/principio activo:
    todo lo que aparece ANTES de la primera palabra de forma farmacéutica.
    """
    pos = len(nombre_norm)
    for kw in FORMAS_FARMACEUTICAS:
        m = re.search(kw, nombre_norm, re.IGNORECASE)
        if m and m.start() < pos:
            pos = m.start()
    return nombre_norm[:pos].strip()


def extraer_concentracion_primaria(nombre_norm: str) -> str:
    """
    Extrae la concentración principal del medicamento.
    Prioridad: combinaciones > ratio > simple.
    Retorna una cadena normalizada para comparación.
    """
    m = PATRON_CONC.search(nombre_norm)
    if not m:
        return ''

    if m.group(1) and m.group(2):
        # Combinación: normaliza los números eliminando espacios alrededor de '+'
        nums = re.sub(r'\s*\+\s*', '+', m.group(1))
        return f"{nums} {m.group(2).upper()}"
    elif m.group(3) and m.group(4) and m.group(5) and m.group(6):
        # Ratio
        num1 = m.group(3).replace(',', '.')
        num2 = m.group(5).replace(',', '.')
        return f"{num1} {m.group(4).upper()}/{num2} {m.group(6).upper()}"
    elif m.group(7) and m.group(8):
        # Simple
        num = m.group(7).replace(',', '.')
        return f"{num} {m.group(8).upper()}"

    return ''


def firma_farmacologica(nombre: str) -> tuple[str, str]:
    """
    Retorna (molecula, concentracion) normalizadas para comparar.
    """
    n = normalizar_nombre(nombre)
    mol = extraer_molecula(n)
    conc = extraer_concentracion_primaria(n)
    return mol, conc


# ─────────────────────────────────────────────────────────────────────────────
# 3. ANÁLISIS DE CONSISTENCIA POR AGRUPADOR
# ─────────────────────────────────────────────────────────────────────────────

print("\nAnalizando consistencia de agrupadores...")

df['_molecula']     = df['nombre'].apply(lambda x: firma_farmacologica(x)[0])
df['_concentracion'] = df['nombre'].apply(lambda x: firma_farmacologica(x)[1])

resultados = []

for agr, grp in df.groupby('agrupador'):
    mols  = grp['_molecula'].unique().tolist()
    concs = grp['_concentracion'].unique().tolist()

    # Un grupo es válido si tiene UNA sola molécula y UNA sola concentración
    valido = (len(mols) == 1 and len(concs) == 1)

    resultados.append({
        'agrupador': agr,
        'n_items': len(grp),
        'valido': valido,
        'moleculas': mols,
        'concentraciones': concs,
    })

res_df = pd.DataFrame(resultados)
validos   = res_df[res_df['valido']]
invalidos = res_df[~res_df['valido']]

print(f"\n{'─'*50}")
print(f"  Agrupadores VÁLIDOS:   {len(validos):>5}")
print(f"  Agrupadores INVÁLIDOS: {len(invalidos):>5}")
print(f"  Total:                 {len(res_df):>5}")
print(f"{'─'*50}")


# ─────────────────────────────────────────────────────────────────────────────
# 4. PROPUESTA DE NUEVOS AGRUPADORES
# ─────────────────────────────────────────────────────────────────────────────

print("\nGenerando propuesta de reclasificación...")

# Número siguiente disponible para nuevos agrupadores
nums_existentes = [
    int(re.search(r'[A-Z](\d+)', a).group(1))
    for a in df['agrupador'].unique()
    if re.search(r'[A-Z](\d+)', str(a))
]
siguiente_num = max(nums_existentes) + 1

# Diccionario de mapeo: índice original → (nuevo_agrupador, nuevo_codigo)
mapa_nuevos = {}   # {idx: {'agrupador_nuevo': ..., 'codigo_nuevo': ...}}

tipo_cambio = {}   # {idx: 'VALIDO' | 'RECLASIFICADO' | 'VALIDO_UNICO'}

# Procesar grupos inválidos
for _, row in invalidos.iterrows():
    agr  = row['agrupador']
    grp  = df[df['agrupador'] == agr].copy()

    # Construir subgrupos por firma (molecula + concentracion)
    grp['_firma'] = grp.apply(
        lambda r: (r['_molecula'], r['_concentracion']), axis=1
    )
    subgrupos = grp.groupby('_firma', sort=False)

    firmas_unicas = list(grp['_firma'].unique())
    n_firmas = len(firmas_unicas)

    for i_firma, firma in enumerate(firmas_unicas):
        sub = grp[grp['_firma'] == firma]
        indices = sub.index.tolist()

        # Si es la primera firma del grupo original, puede conservar el agrupador
        # original SOLO si los ítems de ESE subgrupo son todos los que comparten esa firma
        # y el agrupador original no se usa para otra firma.
        # Para simplificar y máxima claridad: siempre se asigna nuevo código a grupos inválidos.

        nuevo_agr = f"M{siguiente_num:06d}"
        siguiente_num += 1

        # Asignar códigos secuenciales: primer ítem = agrupador raíz,
        # segundo = agrupador+01, tercero = agrupador+02, etc.
        for j, idx in enumerate(indices):
            if j == 0:
                nuevo_cod = nuevo_agr
            else:
                nuevo_cod = f"{nuevo_agr}{j:02d}"

            mapa_nuevos[idx] = {
                'agrupador_nuevo': nuevo_agr,
                'codigo_nuevo': nuevo_cod,
            }
            tipo_cambio[idx] = 'RECLASIFICADO'

# Grupos válidos: conservan agrupador y codigo original
for idx in df.index:
    if idx not in mapa_nuevos:
        agr = df.at[idx, 'agrupador']
        cod = df.at[idx, 'codigo']
        grp_size = df[df['agrupador'] == agr].shape[0]
        mapa_nuevos[idx] = {
            'agrupador_nuevo': agr,
            'codigo_nuevo': cod,
        }
        tipo_cambio[idx] = 'VALIDO'


# ─────────────────────────────────────────────────────────────────────────────
# 5. CONSTRUCCIÓN DEL DATAFRAME DE SALIDA
# ─────────────────────────────────────────────────────────────────────────────

df_salida = df.copy()
df_salida['agrupador_nuevo'] = [mapa_nuevos[i]['agrupador_nuevo'] for i in df.index]
df_salida['codigo_nuevo']    = [mapa_nuevos[i]['codigo_nuevo']    for i in df.index]
df_salida['estado']          = [tipo_cambio[i] for i in df.index]
df_salida['molecula_detectada']     = df['_molecula']
df_salida['concentracion_detectada'] = df['_concentracion']

# Limpiar columnas internas
df_salida = df_salida.drop(columns=['_molecula', '_concentracion'])


# ─────────────────────────────────────────────────────────────────────────────
# 6. ESTADÍSTICAS DETALLADAS
# ─────────────────────────────────────────────────────────────────────────────

reclasificados = df_salida[df_salida['estado'] == 'RECLASIFICADO']
n_reclasificados = len(reclasificados)
n_invalidos_grupos = len(invalidos)

print(f"\nRegistros reclasificados:   {n_reclasificados:,}")
print(f"Grupos inválidos procesados: {n_invalidos_grupos:,}")
print(f"\nNuevos agrupadores generados: {siguiente_num - max(nums_existentes) - 1:,}")
print(f"Próximo agrupador libre: M{siguiente_num:06d}")

# Detalle de los 20 grupos inválidos más grandes
print(f"\n{'─'*70}")
print("TOP 20 AGRUPADORES INVÁLIDOS (por cantidad de ítems):")
print(f"{'─'*70}")
invalidos_detalle = invalidos.nlargest(20, 'n_items')
for _, row in invalidos_detalle.iterrows():
    print(f"\n  Agrupador: {row['agrupador']}  ({row['n_items']} ítems)")
    print(f"  Moléculas detectadas ({len(row['moleculas'])}):")
    for mol in row['moleculas'][:5]:
        print(f"    • {mol[:80]}")
    print(f"  Concentraciones detectadas ({len(row['concentraciones'])}):")
    for conc in row['concentraciones'][:5]:
        print(f"    • {conc}")


# ─────────────────────────────────────────────────────────────────────────────
# 7. GENERACIÓN DEL EXCEL DE SALIDA
# ─────────────────────────────────────────────────────────────────────────────

print(f"\n{'─'*70}")
print("Generando archivo de salida...")

wb = Workbook()

# ── Hoja 1: Catálogo Reclasificado ──────────────────────────────────────────
ws_cat = wb.active
ws_cat.title = "Catalogo Reclasificado"

# Estilos
FILL_HEADER   = PatternFill("solid", fgColor="1F4E79")
FILL_VALIDO   = PatternFill("solid", fgColor="E8F5E9")
FILL_RECLASIF = PatternFill("solid", fgColor="FFF9C4")
FONT_HEADER   = Font(color="FFFFFF", bold=True, size=10)
FONT_NORMAL   = Font(size=10)
FONT_BOLD     = Font(bold=True, size=10)
ALIGN_CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT    = Alignment(horizontal="left", vertical="center", wrap_text=True)

thin = Side(style="thin", color="BBBBBB")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# Columnas de salida y sus anchos
COLUMNAS = [
    ("agrupador",              "Agrupador\nOriginal",         15),
    ("codigo",                 "Código\nOriginal",            18),
    ("nombre",                 "Nombre del Medicamento",      55),
    ("codigo.1",               "Código Alt.",                 18),
    ("marca",                  "Marca",                       20),
    ("atc",                    "ATC",                         12),
    ("consecutivo_forma",      "Consec.\nForma",               9),
    ("cums",                   "CUMS",                        16),
    ("habilita",               "Habilitado",                  11),
    ("molecula_detectada",     "Molécula\nDetectada",         35),
    ("concentracion_detectada","Concentración\nDetectada",    20),
    ("agrupador_nuevo",        "Agrupador\nNUEVO",            15),
    ("codigo_nuevo",           "Código\nNUEVO",               18),
    ("estado",                 "Estado",                      14),
]

# Cabecera
for col_i, (_, titulo, ancho) in enumerate(COLUMNAS, start=1):
    cell = ws_cat.cell(row=1, column=col_i, value=titulo)
    cell.fill   = FILL_HEADER
    cell.font   = FONT_HEADER
    cell.alignment = ALIGN_CENTER
    cell.border = BORDER
    ws_cat.column_dimensions[get_column_letter(col_i)].width = ancho

ws_cat.row_dimensions[1].height = 30
ws_cat.freeze_panes = "A2"

# Ordenar: primero inválidos reclasificados, luego válidos
df_orden = pd.concat([
    df_salida[df_salida['estado'] == 'RECLASIFICADO'].sort_values('agrupador_nuevo'),
    df_salida[df_salida['estado'] == 'VALIDO'].sort_values('agrupador'),
]).reset_index(drop=True)

# Datos
for row_i, (_, fila) in enumerate(df_orden.iterrows(), start=2):
    fill = FILL_RECLASIF if fila['estado'] == 'RECLASIFICADO' else FILL_VALIDO
    for col_i, (col_name, _, _) in enumerate(COLUMNAS, start=1):
        val = fila.get(col_name, '')
        if pd.isna(val):
            val = ''
        elif isinstance(val, bool):
            val = 'SÍ' if val else 'NO'
        cell = ws_cat.cell(row=row_i, column=col_i, value=str(val) if val != '' else val)
        cell.fill      = fill
        cell.font      = FONT_NORMAL
        cell.alignment = ALIGN_LEFT
        cell.border    = BORDER
    # Resaltar columnas nuevas
    for col_extra in [12, 13, 14]:  # agrupador_nuevo, codigo_nuevo, estado
        ws_cat.cell(row=row_i, column=col_extra).font = FONT_BOLD

# Auto-filter
ws_cat.auto_filter.ref = ws_cat.dimensions


# ── Hoja 2: Resumen de Grupos Inválidos ──────────────────────────────────────
ws_inv = wb.create_sheet("Grupos Invalidos")

FILL_INV_H  = PatternFill("solid", fgColor="C62828")
FILL_INV_R  = PatternFill("solid", fgColor="FFEBEE")
FILL_INV_R2 = PatternFill("solid", fgColor="FFCDD2")

inv_cols = [
    ("Agrupador\nOriginal",       14),
    ("Nº\nÍtems",                  8),
    ("Moléculas\nDetectadas",      50),
    ("Concentraciones\nDetectadas",30),
    ("Agrupadores\nNuevos Propuestos",35),
    ("Tipo de\nError",             20),
]

for col_i, (titulo, ancho) in enumerate(inv_cols, start=1):
    cell = ws_inv.cell(row=1, column=col_i, value=titulo)
    cell.fill      = FILL_INV_H
    cell.font      = Font(color="FFFFFF", bold=True, size=10)
    cell.alignment = ALIGN_CENTER
    cell.border    = BORDER
    ws_inv.column_dimensions[get_column_letter(col_i)].width = ancho

ws_inv.row_dimensions[1].height = 30
ws_inv.freeze_panes = "A2"

# Mapear agrupadores originales → nuevos agrupadores
agr_to_nuevos = df_salida[df_salida['estado'] == 'RECLASIFICADO'] \
    .groupby('agrupador')['agrupador_nuevo'].apply(lambda x: sorted(x.unique().tolist())).to_dict()

for row_i, (_, row) in enumerate(invalidos.sort_values('n_items', ascending=False).iterrows(), start=2):
    agr   = row['agrupador']
    mols  = row['moleculas']
    concs = row['concentraciones']
    nuevos = agr_to_nuevos.get(agr, [])

    # Determinar tipo de error
    if len(mols) > 1:
        tipo = "Mezcla de moléculas"
        if len(concs) > 1:
            tipo = "Mezcla de moléculas y concentraciones"
    else:
        tipo = "Mezcla de concentraciones"

    fill = FILL_INV_R if row_i % 2 == 0 else FILL_INV_R2

    valores = [
        agr,
        row['n_items'],
        " | ".join(str(m)[:60] for m in mols),
        " | ".join(str(c) for c in concs),
        " | ".join(nuevos),
        tipo,
    ]
    for col_i, val in enumerate(valores, start=1):
        cell = ws_inv.cell(row=row_i, column=col_i, value=val)
        cell.fill      = fill
        cell.font      = FONT_NORMAL
        cell.alignment = ALIGN_LEFT
        cell.border    = BORDER

ws_inv.auto_filter.ref = ws_inv.dimensions


# ── Hoja 3: Resumen Estadístico ──────────────────────────────────────────────
ws_res = wb.create_sheet("Resumen")

FILL_STAT_H = PatternFill("solid", fgColor="004D40")
FILL_STAT_V = PatternFill("solid", fgColor="E0F2F1")
FILL_STAT_I = PatternFill("solid", fgColor="FFF8E1")
FILL_STAT_N = PatternFill("solid", fgColor="E8EAF6")

stats = [
    ("ESTADÍSTICAS GENERALES", None, FILL_STAT_H, True),
    ("Total de registros en el catálogo",          f"{len(df):,}",              FILL_STAT_V, False),
    ("Total de agrupadores únicos originales",      f"{df['agrupador'].nunique():,}", FILL_STAT_V, False),
    ("", "", PatternFill(), False),
    ("RESULTADO DEL ANÁLISIS", None, FILL_STAT_H, True),
    ("Agrupadores VÁLIDOS (sin cambios)",           f"{len(validos):,}",         FILL_STAT_V, False),
    ("Agrupadores INVÁLIDOS (requieren corrección)",f"{len(invalidos):,}",       FILL_STAT_I, False),
    ("Registros que cambian de agrupador",          f"{n_reclasificados:,}",     FILL_STAT_I, False),
    ("Registros sin cambios",                       f"{len(df) - n_reclasificados:,}", FILL_STAT_V, False),
    ("", "", PatternFill(), False),
    ("NUEVOS AGRUPADORES", None, FILL_STAT_H, True),
    ("Nuevos agrupadores generados",               f"{siguiente_num - max(nums_existentes) - 1:,}", FILL_STAT_N, False),
    ("Rango de nuevos agrupadores",                f"M{max(nums_existentes)+1:06d} – M{siguiente_num-1:06d}", FILL_STAT_N, False),
    ("Próximo agrupador libre",                    f"M{siguiente_num:06d}",      FILL_STAT_N, False),
    ("", "", PatternFill(), False),
    ("TIPOS DE ERROR EN GRUPOS INVÁLIDOS", None, FILL_STAT_H, True),
]

# Contar tipos de error
solo_mol  = sum(1 for _, r in invalidos.iterrows() if len(r['moleculas']) > 1 and len(r['concentraciones']) == 1)
solo_conc = sum(1 for _, r in invalidos.iterrows() if len(r['moleculas']) == 1 and len(r['concentraciones']) > 1)
ambos     = sum(1 for _, r in invalidos.iterrows() if len(r['moleculas']) > 1 and len(r['concentraciones']) > 1)

stats += [
    ("Solo mezcla de moléculas diferentes",         f"{solo_mol:,}",    FILL_STAT_I, False),
    ("Solo mezcla de concentraciones diferentes",   f"{solo_conc:,}",   FILL_STAT_I, False),
    ("Mezcla de moléculas y concentraciones",        f"{ambos:,}",       FILL_STAT_I, False),
]

ws_res.column_dimensions['A'].width = 46
ws_res.column_dimensions['B'].width = 26

for row_i, entry in enumerate(stats, start=1):
    label, valor, fill, is_title = entry
    ws_res.row_dimensions[row_i].height = 22

    cell_a = ws_res.cell(row=row_i, column=1, value=label)
    cell_a.fill = fill
    cell_a.border = BORDER

    if is_title:
        cell_a.font      = Font(color="FFFFFF", bold=True, size=11)
        cell_a.alignment = ALIGN_CENTER
        ws_res.merge_cells(f"A{row_i}:B{row_i}")
    else:
        cell_a.font      = FONT_NORMAL
        cell_a.alignment = ALIGN_LEFT

        cell_b = ws_res.cell(row=row_i, column=2, value=valor)
        cell_b.fill      = fill
        cell_b.font      = FONT_BOLD
        cell_b.alignment = ALIGN_CENTER
        cell_b.border    = BORDER


# ─────────────────────────────────────────────────────────────────────────────
# 8. GUARDAR
# ─────────────────────────────────────────────────────────────────────────────
wb.save(RUTA_SALIDA)

print(f"\n{'='*70}")
print(f"✓ Archivo generado exitosamente:")
print(f"  {RUTA_SALIDA}")
print(f"\nContenido del archivo:")
print(f"  Hoja 1 - 'Catalogo Reclasificado': {len(df_orden):,} registros")
print(f"  Hoja 2 - 'Grupos Invalidos':        {len(invalidos):,} grupos con inconsistencias")
print(f"  Hoja 3 - 'Resumen':                 Estadísticas del proceso")
print(f"{'='*70}")
